from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from .calendar import TOTAL_SEMANAS, es_apto, format_antiguedad, parse_iso_date, vacation_record_for
from .plan import group_periods


def programmed_dnis(targets, daily_set) -> set[str]:
    dnis = {str(dni) for (dni, _week), dias in (targets or {}).items() if int(dias) > 0}
    if daily_set:
        dnis.update(item.split("|", 1)[0] for item in daily_set if "|" in item)
    return dnis


def filter_employees_for_export(
    employees,
    targets,
    daily_set,
    today: date | None = None,
    *,
    solo_aptos: bool = False,
    con_vacaciones: bool = False,
):
    have = programmed_dnis(targets, daily_set) if con_vacaciones else None
    return [
        e
        for e in employees
        if (not solo_aptos or es_apto(e, today)) and (have is None or str(e["dni"]) in have)
    ]


def semaforo_fill(value):
    try:
        value = int(value)
    except Exception:
        return None
    colors = {1: "F8696B", 2: "F8696B", 3: "F9C47A", 4: "FFEB84", 5: "FFEB84", 6: "9ED17A", 7: "63BE7B"}
    color = colors.get(value)
    return PatternFill("solid", fgColor=color) if color else None


def build_record(employees, dias_map: dict[str, list[date]], year: int, today: date | None = None):
    rows = []
    for w in employees:
        dni = str(w["dni"])
        f_ingreso = parse_iso_date(w.get("fecha_ingreso"))
        rec = vacation_record_for(f_ingreso, dias_map.get(dni, []), year, today)
        cumple = rec["cumple_record"]
        vencimiento = rec["fecha_vencimiento"]
        rows.append({
            "EMPRESA": w.get("empresa", ""),
            "DNI": dni,
            "NOMBRE": w["nombre"],
            "FECHA_INGRESO": f_ingreso.strftime("%d/%m/%Y") if f_ingreso else "",
            "DIVISION": w.get("division", ""),
            "GERENCIA": w["gerencia"],
            "AREA": w["area"],
            "RECORD_VACACIONAL": rec["record_vacacional"],
            "CUMPLE_RECORD": cumple.strftime("%d/%m/%Y") if cumple else "",
            "DIAS_PROGRAMADOS": rec["dias_programados"],
            "DIAS_GOZADOS": rec["dias_gozados"],
            "DIAS_PENDIENTES": rec["dias_pendientes"],
            "FECHA_VENCIMIENTO": vencimiento.strftime("%d/%m/%Y") if vencimiento else "",
            "ANTIGUEDAD": format_antiguedad(f_ingreso),
        })
    return rows


ALL_SHEETS = {"RESUMEN", "PLANIFICACION", "PERIODOS", "RECORD_VACACIONAL", "CAMBIOS"}
RECORD_SHEETS = {"RECORD_VACACIONAL"}


def _style_workbook(wb, *, year, current_year, current_week):
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="0F172A")
    current_fill = PatternFill("solid", fgColor="0F766E")
    history_fill = PatternFill("solid", fgColor="D0D5DD")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            max_len = 0
            for cell in column_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
                cell.border = border
                cell.alignment = Alignment(vertical="center")
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 10), 35)

    if "PLANIFICACION" not in wb.sheetnames:
        return

    ws = wb["PLANIFICACION"]
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col_idx).value or "")
        if header.startswith("S") and header[1:].isdigit():
            week = int(header[1:])
            if year == current_year and week == current_week:
                ws.cell(1, col_idx).fill = current_fill
            elif year == current_year and week < current_week:
                ws.cell(1, col_idx).fill = history_fill
            for row_idx in range(2, ws.max_row + 1):
                fill = semaforo_fill(ws.cell(row_idx, col_idx).value)
                if fill:
                    ws.cell(row_idx, col_idx).fill = fill
                    ws.cell(row_idx, col_idx).font = Font(bold=True, color="000000")

    legend = [
        ("SEMAFORO 1–7 DÍAS", ""),
        ("1–2 días", "F8696B"),
        ("3 días", "F9C47A"),
        ("4–5 días", "FFEB84"),
        ("6 días", "9ED17A"),
        ("7 días", "63BE7B"),
        ("0 días", "FFFFFF"),
    ]
    start_col = ws.max_column + 2
    for i, (label, color) in enumerate(legend, start=1):
        c1 = ws.cell(1 + i, start_col, label)
        if color:
            c1.fill = PatternFill("solid", fgColor=color)
        c1.border = border
    ws.freeze_panes = "F2"


def export_excel(
    employees,
    daily_set,
    targets,
    year,
    label_jefatura,
    current_year,
    current_week,
    change_log,
    usuario,
    nombre_usuario,
    historial=None,
    sheets: set[str] | None = None,
):
    wanted = sheets or ALL_SHEETS
    frames: list[tuple[str, pd.DataFrame]] = []

    if "RESUMEN" in wanted or "PERIODOS" in wanted:
        periods = group_periods(employees, daily_set, year)
    else:
        periods = []

    if "RESUMEN" in wanted:
        frames.append(("RESUMEN", pd.DataFrame([{
            "JEFATURA": label_jefatura,
            "AÑO": year,
            "TRABAJADORES": len(employees),
            "PERSONAS_PROGRAMADAS": len({r["dni"] for r in periods}),
            "DIAS_PROGRAMADOS": sum(int(r["dias"]) for r in periods),
            "GENERADO": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            "USUARIO_GENERADOR": usuario,
            "NOMBRE_USUARIO": nombre_usuario,
        }])))

    if "PLANIFICACION" in wanted:
        weekly_rows = []
        for w in sorted(employees, key=lambda x: str(x["nombre"]).casefold()):
            dni = str(w["dni"])
            weeks = {f"S{week}": int(targets.get((dni, week), 0)) for week in range(1, TOTAL_SEMANAS + 1)}
            weekly_rows.append({
                "NOMBRE": w["nombre"],
                "DNI": dni,
                "AREA": w["area"],
                "TIPO": w["tipo_personal"],
                "TOTAL": sum(weeks.values()),
                **weeks,
            })
        frames.append(("PLANIFICACION", pd.DataFrame(weekly_rows)))

    if "PERIODOS" in wanted:
        periods_df = pd.DataFrame(periods) if periods else pd.DataFrame(
            columns=["dni", "nombre", "gerencia", "area", "jefatura", "tipo_personal", "fecha_inicio", "fecha_fin", "dias"]
        )
        frames.append(("PERIODOS", periods_df))

    if "RECORD_VACACIONAL" in wanted:
        record_df = pd.DataFrame(historial) if historial else pd.DataFrame()
        frames.append(("RECORD_VACACIONAL", record_df))

    if "CAMBIOS" in wanted:
        frames.append(("CAMBIOS", pd.DataFrame(change_log) if change_log else pd.DataFrame()))

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in frames:
            df.to_excel(writer, sheet_name=name, index=False)
    output.seek(0)

    wb = load_workbook(output)
    _style_workbook(wb, year=year, current_year=current_year, current_week=current_week)
    output2 = BytesIO()
    wb.save(output2)
    output2.seek(0)
    return output2.getvalue()
