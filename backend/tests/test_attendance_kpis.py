from datetime import date, time

from app.domain.attendance_kpis import (
    JORNADA_MINUTOS,
    build_asistencia_kpis,
    classify_shift,
    lost_minutes,
)


def test_margen_0650_1705_cumple():
    assert classify_shift(time(6, 50), time(17, 5), 2) == "cumple"
    assert lost_minutes("cumple", time(6, 50), time(17, 5)) == 0


def test_oficial_0630_1705_cumple():
    assert classify_shift(time(6, 30), time(17, 5), 2) == "cumple"


def test_bus_0650_1720_cumple():
    assert classify_shift(time(6, 50, 40), time(17, 20), 2) == "cumple"


def test_sale_1605_no_cumple():
    assert classify_shift(time(6, 30), time(16, 5), 2) == "salida_temprano"
    assert lost_minutes("salida_temprano", time(6, 30), time(16, 5)) == 60


def test_llega_0700_tardanza():
    assert classify_shift(time(7, 0), time(17, 5), 2) == "tardanza"
    assert lost_minutes("tardanza", time(7, 0), time(17, 5)) == 10


def test_una_marca_es_no_marcan():
    assert classify_shift(time(6, 40), time(6, 40), 1) == "no_marcan"
    assert lost_minutes("no_marcan", None, None) == JORNADA_MINUTOS


def test_kpi_agrega_personas_fuera_de_margen():
    emp = {
        "dni": "1",
        "nombre": "Ana",
        "area": "TI",
        "cargo_actual": "Analista",
        "tipo_personal": "ADMINISTRATIVO",
        "fecha_ingreso": "2020-01-01",
    }
    # miércoles 2026-08-19
    d = date(2026, 8, 19)
    shifts = {
        "1": {
            d: {"entrada": time(8, 0), "salida": time(15, 0), "n": 2},
        }
    }
    out = build_asistencia_kpis(
        [emp],
        set(),
        shifts,
        start=d,
        end=d,
        configured=True,
        times_ok=True,
    )
    assert out["personas_incumplen"] == 1
    assert out["llegadas_tarde"]["casos"] == 1
    assert out["salidas_temprano"]["casos"] == 1
    assert out["incumplimiento_pct"] == 100.0
    assert out["detalle_incumplimiento"]["jornadas_con_2_marcas"] == 1
    assert out["detalle_incumplimiento"]["jornadas_fuera"] == 1
    assert out["detalle_incumplimiento"]["ambos"] == 1
    assert "S/" not in (out.get("alerta") or {}).get("texto", "")


def test_no_marcan_no_infla_incumplimiento_de_jornada():
    emp = {
        "dni": "1",
        "nombre": "Ana",
        "area": "TI",
        "cargo_actual": "Analista",
        "tipo_personal": "ADMINISTRATIVO",
        "fecha_ingreso": "2020-01-01",
    }
    ok = date(2026, 8, 19)
    falta = date(2026, 8, 20)
    shifts = {
        "1": {
            ok: {"entrada": time(6, 50), "salida": time(17, 10), "n": 2},
        }
    }
    out = build_asistencia_kpis(
        [emp], set(), shifts, start=ok, end=falta, configured=True, times_ok=True
    )
    assert out["incumplimiento_pct"] == 0.0
    assert out["cumplimiento_jornada_pct"] == 100.0
    assert out["no_marcan"]["casos"] == 1
    assert out["no_marcan"]["personas"] == 1
    assert out["no_marcan"]["sin_marca"] == 1
    assert out["no_marcan"]["por_dia"][0]["fecha"] == "2026-08-20"
    assert out["personas_incumplen"] == 0


def test_trujillo_0800_1800_cumple():
    assert (
        classify_shift(
            time(8, 0),
            time(18, 0),
            2,
            entrada_limite=time(8, 20),
            salida_minima=time(18, 0),
        )
        == "cumple"
    )


def test_tesoreria_trujillo_no_es_100_por_horario_oficina():
    emp = {
        "dni": "1",
        "nombre": "Ana",
        "area": "TESORERIA",
        "cargo_actual": "Analista",
        "tipo_personal": "ADMINISTRATIVO",
        "fecha_ingreso": "2020-01-01",
        "gerencia": "Administración",
    }
    d = date(2026, 8, 19)
    shifts = {
        "1": {
            d: {
                "entrada": time(8, 5),
                "salida": time(18, 10),
                "n": 2,
                "dispositivo": "TRUJILLO - EL GOLF",
            }
        }
    }
    out = build_asistencia_kpis(
        [emp], set(), shifts, start=d, end=d, configured=True, times_ok=True
    )
    assert out["incumplimiento_pct"] == 0.0
    tes = next(r for r in out["ranking_area"] if r["name"] == "TESORERIA")
    assert tes["pct"] == 0.0
    assert "al menos un incumplimiento" not in " ".join(r["body"] for r in out["recomendaciones"])
    emp = {
        "dni": "1",
        "nombre": "Luis",
        "area": "Campo",
        "cargo_actual": "Cosechador",
        "tipo_personal": "OPERATIVO",
        "fecha_ingreso": "2020-01-01",
    }
    d = date(2026, 8, 19)
    shifts = {"1": {d: {"entrada": time(8, 0), "salida": time(15, 0), "n": 2}}}
    out = build_asistencia_kpis(
        [emp], set(), shifts, start=d, end=d, configured=True, times_ok=True
    )
    assert out["incumplimiento_pct"] == 0.0
    assert out["llegadas_tarde"]["casos"] == 0


def test_presencia_sin_hora_no_es_incumplimiento():
    emp = {
        "dni": "1",
        "nombre": "Ana",
        "area": "TI",
        "cargo_actual": "Analista",
        "tipo_personal": "ADMINISTRATIVO",
        "fecha_ingreso": "2020-01-01",
    }
    d = date(2026, 8, 19)
    shifts = {"1": {d: {"entrada": None, "salida": None, "n": 0, "n_rows": 3}}}
    out = build_asistencia_kpis(
        [emp], set(), shifts, start=d, end=d, configured=True, times_ok=False
    )
    assert out["personas_incumplen"] == 0
    assert out["no_marcan"]["casos"] == 0


def test_una_marca_cuenta_aparte_en_no_marcan():
    emp = {
        "dni": "1",
        "nombre": "Ana",
        "area": "TI",
        "cargo_actual": "Analista",
        "tipo_personal": "ADMINISTRATIVO",
        "fecha_ingreso": "2020-01-01",
    }
    d = date(2026, 8, 19)
    shifts = {"1": {d: {"entrada": time(6, 40), "salida": time(6, 40), "n": 1}}}
    out = build_asistencia_kpis([emp], set(), shifts, start=d, end=d, configured=True, times_ok=True)
    assert out["no_marcan"]["casos"] == 1
    assert out["no_marcan"]["una_marca"] == 1
    assert out["no_marcan"]["sin_marca"] == 0


def test_operativo_sin_marca_no_es_oficina():
    emp = {
        "dni": "1",
        "nombre": "Luis",
        "area": "Campo",
        "cargo_actual": "Cosechador",
        "tipo_personal": "OPERATIVO",
        "fecha_ingreso": "2020-01-01",
    }
    d = date(2026, 8, 19)
    out = build_asistencia_kpis([emp], set(), {}, start=d, end=d, configured=True, times_ok=True)
    assert out["no_marcan"]["casos"] == 1
    assert out["no_marcan"]["operativo_sin_marca"] == 1


def test_excel_asistencia_tiene_hojas():
    from io import BytesIO

    from openpyxl import load_workbook

    from app.domain.attendance_export import export_asistencia_excel

    emp = {
        "dni": "1",
        "nombre": "Ana",
        "area": "TI",
        "cargo_actual": "Analista",
        "tipo_personal": "ADMINISTRATIVO",
        "fecha_ingreso": "2020-01-01",
    }
    d = date(2026, 8, 19)
    kpis = build_asistencia_kpis([emp], set(), {}, start=d, end=d, configured=True, times_ok=True)
    raw = export_asistencia_excel([emp], set(), {}, start=d, end=d, kpis=kpis)
    wb = load_workbook(BytesIO(raw))
    assert "No_marcan" in wb.sheetnames
    assert "No_marcan_por_dia" in wb.sheetnames
    assert "Detalle_diario" in wb.sheetnames
    assert wb["No_marcan"].max_row >= 2


def test_cesado_sin_fecha_no_infla_no_marcan():
    emp = {
        "dni": "1",
        "nombre": "Ana",
        "area": "TI",
        "cargo_actual": "Analista",
        "tipo_personal": "ADMINISTRATIVO",
        "fecha_ingreso": "2020-01-01",
        "vigencia": "CESADO",
    }
    d = date(2026, 8, 19)
    out = build_asistencia_kpis([emp], set(), {}, start=d, end=d, configured=True, times_ok=True)
    assert out["evaluados"] == 0
    assert out["no_marcan"]["casos"] == 0


def test_fecha_cese_no_cuenta_dias_siguientes():
    emp = {
        "dni": "1",
        "nombre": "Ana",
        "area": "TI",
        "cargo_actual": "Analista",
        "tipo_personal": "ADMINISTRATIVO",
        "fecha_ingreso": "2020-01-01",
        "fecha_cese": "2026-08-19",
    }
    out = build_asistencia_kpis(
        [emp], set(), {}, start=date(2026, 8, 19), end=date(2026, 8, 20), configured=True, times_ok=True
    )
    assert out["no_marcan"]["casos"] == 1
    assert out["no_marcan"]["por_dia"][0]["fecha"] == "2026-08-19"
