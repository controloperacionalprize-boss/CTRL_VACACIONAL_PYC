from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app.domain.calendar import vacation_record_for
from app.domain.employee_calendar import employee_calendar_payload
from app.domain.export import export_excel, filter_employees_for_export


def test_record_primer_anio():
    rec = vacation_record_for(date(2026, 1, 1), [], 2026, today=date(2026, 8, 21))
    assert rec["record_vacacional"] == "2026-2027"
    assert rec["cumple_record"] == date(2027, 1, 1)
    assert rec["record_cumplido"] is False
    assert rec["dias_pendientes"] == 30


def test_record_cuenta_programados_y_gozados():
    programmed = [date(2026, 8, 1), date(2026, 8, 2), date(2026, 9, 1)]
    rec = vacation_record_for(date(2026, 1, 1), programmed, 2026, today=date(2026, 8, 21))
    assert rec["dias_programados"] == 3
    assert rec["dias_gozados"] == 2  # 1 y 2 ago ya pasaron
    assert rec["dias_pendientes"] == 27


def test_calendar_payload_incluye_record():
    worker = {
        "dni": "1",
        "nombre": "Pepe",
        "empresa": "X",
        "area": "A",
        "cargo_actual": "C",
        "fecha_ingreso": "2026-01-01",
        "gerencia": "G",
    }
    payload = employee_calendar_payload(worker, 2026, [date(2026, 9, 1), date(2026, 9, 2)])
    assert payload["record"]["record_vacacional"] == "2026-2027"
    assert payload["record"]["dias_programados"] == 2
    assert payload["periodos"][0]["inicio"] == "2026-09-01"
    assert payload["periodos"][0]["fin"] == "2026-09-02"


def test_excel_hojas_alineadas_con_web():
    data = export_excel([], set(), {}, 2026, "X", 2026, 34, [], "u", "n", [])
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["RESUMEN", "PLANIFICACION", "PERIODOS", "RECORD_VACACIONAL", "CAMBIOS"]
    assert "DETALLE_DIARIO" not in wb.sheetnames
    assert "PLAN_SEMANAL" not in wb.sheetnames
    assert "VACACIONES" not in wb.sheetnames


def test_excel_aptos_solo_record():
    hist = [{
        "EMPRESA": "X",
        "DNI": "1",
        "NOMBRE": "Apto",
        "FECHA_INGRESO": "01/01/2024",
        "DIVISION": "",
        "GERENCIA": "G",
        "AREA": "A",
        "RECORD_VACACIONAL": "2025-2026",
        "CUMPLE_RECORD": "01/01/2025",
        "DIAS_PROGRAMADOS": 15,
        "DIAS_GOZADOS": 0,
        "DIAS_PENDIENTES": 15,
        "FECHA_VENCIMIENTO": "31/12/2025",
        "ANTIGUEDAD": "2 años",
    }]
    data = export_excel([], set(), {}, 2026, "APTOS", 2026, 34, [], "u", "n", hist, sheets={"RECORD_VACACIONAL"})
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["RECORD_VACACIONAL"]
    assert wb["RECORD_VACACIONAL"]["C2"].value == "Apto"


def test_filtra_aptos_con_vacaciones():
    today = date(2026, 8, 31)
    apto = {
        "dni": "1",
        "nombre": "Apto",
        "fecha_ingreso": "2024-01-01",
        "area": "A",
        "tipo_personal": "ADM",
        "gerencia": "G",
    }
    adelanto = {**apto, "dni": "2", "nombre": "Nuevo", "fecha_ingreso": "2026-03-01"}
    apto_sin = {**apto, "dni": "3", "nombre": "Sin días"}
    out = filter_employees_for_export(
        [apto, adelanto, apto_sin],
        {("1", 10): 5, ("2", 10): 3},
        set(),
        today,
        solo_aptos=True,
        con_vacaciones=True,
    )
    assert [e["dni"] for e in out] == ["1"]
